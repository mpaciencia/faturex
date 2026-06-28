"""
Rota de relatórios — GET /api/relatorios.

Gera ficheiro .xlsx com openpyxl, organizado por Despesas e Receitas.
"""

import io
import logging
from datetime import date
from decimal import Decimal
from urllib.parse import unquote
from zipfile import ZIP_DEFLATED, ZipFile

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from services import supabase_client

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/relatorios", tags=["Relatórios"])


def _build_sheet(
    wb: Workbook,
    sheet_name: str,
    faturas: list[dict],
) -> None:
    """
    Constrói uma folha no workbook com faturas agrupadas por categoria,
    incluindo subtotais por categoria e total geral.

    Colunas: Nome Empresa | Data | NIF | Categoria | Valor IVA | Valor Total | Observações
    """
    ws = wb.create_sheet(title=sheet_name)

    # --- Estilos ---
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    subtotal_font = Font(name="Calibri", bold=True, size=10)
    subtotal_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    total_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    total_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # --- Headers ---
    headers = [
        "Nome Empresa",
        "Data",
        "NIF",
        "Categoria",
        "Valor IVA",
        "Valor Total",
        "Observações",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # --- Agrupar faturas por categoria ---
    por_categoria: dict[str, list[dict]] = {}
    for fatura in faturas:
        cat = fatura.get("categoria", "Outros")
        por_categoria.setdefault(cat, []).append(fatura)

    # Ordenar categorias alfabeticamente
    categorias_ordenadas = sorted(por_categoria.keys())

    row_idx = 2
    total_geral_iva = Decimal("0")
    total_geral_valor = Decimal("0")

    for categoria in categorias_ordenadas:
        faturas_cat = por_categoria[categoria]
        subtotal_iva = Decimal("0")
        subtotal_valor = Decimal("0")

        for fatura in faturas_cat:
            iva = Decimal(str(fatura.get("imposto_total", "0")))
            total = Decimal(str(fatura.get("valor_total", "0")))

            ws.cell(row=row_idx, column=1, value=fatura.get("nome_emissor", "")).border = thin_border
            ws.cell(row=row_idx, column=2, value=fatura.get("data_fatura", "")).border = thin_border
            ws.cell(row=row_idx, column=3, value=fatura.get("nif_emissor", "")).border = thin_border
            ws.cell(row=row_idx, column=4, value=categoria).border = thin_border

            cell_iva = ws.cell(row=row_idx, column=5, value=float(iva))
            cell_iva.number_format = '#,##0.00'
            cell_iva.border = thin_border

            cell_total = ws.cell(row=row_idx, column=6, value=float(total))
            cell_total.number_format = '#,##0.00'
            cell_total.border = thin_border

            ws.cell(row=row_idx, column=7, value=fatura.get("observacoes") or "").border = thin_border

            subtotal_iva += iva
            subtotal_valor += total
            row_idx += 1

        # Linha de subtotal da categoria
        ws.cell(row=row_idx, column=1, value="").border = thin_border
        ws.cell(row=row_idx, column=2, value="").border = thin_border
        ws.cell(row=row_idx, column=3, value="").border = thin_border

        cell_sub_label = ws.cell(row=row_idx, column=4, value=f"Subtotal — {categoria}")
        cell_sub_label.font = subtotal_font
        cell_sub_label.fill = subtotal_fill
        cell_sub_label.border = thin_border

        cell_sub_iva = ws.cell(row=row_idx, column=5, value=float(subtotal_iva))
        cell_sub_iva.number_format = '#,##0.00'
        cell_sub_iva.font = subtotal_font
        cell_sub_iva.fill = subtotal_fill
        cell_sub_iva.border = thin_border

        cell_sub_total = ws.cell(row=row_idx, column=6, value=float(subtotal_valor))
        cell_sub_total.number_format = '#,##0.00'
        cell_sub_total.font = subtotal_font
        cell_sub_total.fill = subtotal_fill
        cell_sub_total.border = thin_border

        ws.cell(row=row_idx, column=7, value="").border = thin_border

        total_geral_iva += subtotal_iva
        total_geral_valor += subtotal_valor
        row_idx += 1

    # --- Linha de total geral ---
    ws.cell(row=row_idx, column=1, value="").border = thin_border
    ws.cell(row=row_idx, column=2, value="").border = thin_border
    ws.cell(row=row_idx, column=3, value="").border = thin_border

    cell_total_label = ws.cell(row=row_idx, column=4, value="TOTAL GERAL")
    cell_total_label.font = total_font
    cell_total_label.fill = total_fill
    cell_total_label.border = thin_border

    cell_total_iva = ws.cell(row=row_idx, column=5, value=float(total_geral_iva))
    cell_total_iva.number_format = '#,##0.00'
    cell_total_iva.font = total_font
    cell_total_iva.fill = total_fill
    cell_total_iva.border = thin_border

    cell_total_val = ws.cell(row=row_idx, column=6, value=float(total_geral_valor))
    cell_total_val.number_format = '#,##0.00'
    cell_total_val.font = total_font
    cell_total_val.fill = total_fill
    cell_total_val.border = thin_border

    ws.cell(row=row_idx, column=7, value="").border = thin_border

    # --- Ajustar largura das colunas ---
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 30


def _build_relatorio_response(data_inicio: date, data_fim: date) -> StreamingResponse:
    if data_inicio > data_fim:
        raise HTTPException(
            status_code=400,
            detail="'data_inicio' não pode ser posterior a 'data_fim'.",
        )

    try:
        faturas = supabase_client.get_faturas_by_period(
            data_inicio=data_inicio.isoformat(),
            data_fim=data_fim.isoformat(),
        )
    except Exception:
        logger.exception("Erro ao consultar faturas para relatório Excel")
        raise HTTPException(
            status_code=500,
            detail="Erro interno ao consultar faturas.",
        )

    despesas = [f for f in faturas if f.get("tipo") == "Despesa"]
    receitas = [f for f in faturas if f.get("tipo") == "Receita"]

    wb = Workbook()
    wb.remove(wb.active)

    _build_sheet(wb, "Despesas", despesas)
    _build_sheet(wb, "Receitas", receitas)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"relatorio_{data_inicio.strftime('%Y%m%d')}_{data_fim.strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        content=buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


@router.get("")
@router.get("/excel")
async def gerar_relatorio(
    data_inicio: date = Query(..., description="Data de início (YYYY-MM-DD)"),
    data_fim: date = Query(..., description="Data de fim (YYYY-MM-DD)"),
):
    logger.info("Recebida requisição GET /excel para gerar relatório no período: %s a %s", data_inicio, data_fim)
    """
    Fluxo C — Geração de relatório Excel.

    Query params: data_inicio e data_fim no formato YYYY-MM-DD.
    Devolve ficheiro .xlsx com folhas 'Despesas' e 'Receitas'.
    """
    return _build_relatorio_response(data_inicio, data_fim)


def _build_zip_response(data_inicio: date, data_fim: date) -> StreamingResponse:
    if data_inicio > data_fim:
        raise HTTPException(
            status_code=400,
            detail="'data_inicio' não pode ser posterior a 'data_fim'.",
        )

    try:
        faturas = supabase_client.get_faturas_by_period(
            data_inicio=data_inicio.isoformat(),
            data_fim=data_fim.isoformat(),
        )
    except Exception:
        logger.exception("Erro ao consultar faturas para arquivo ZIP")
        raise HTTPException(
            status_code=500,
            detail="Erro interno ao consultar faturas.",
        )

    zip_buffer = io.BytesIO()

    try:
        with ZipFile(zip_buffer, mode="w", compression=ZIP_DEFLATED) as archive:
            for index, fatura in enumerate(faturas, start=1):
                url_documento = fatura.get("url_documento")
                if not isinstance(url_documento, str) or not url_documento.strip():
                    raise HTTPException(
                        status_code=500,
                        detail="Fatura sem URL de documento associada.",
                    )

                storage_path = supabase_client.storage_path_from_public_url(url_documento)
                documento_bytes = supabase_client.download_documento(storage_path)

                nome_arquivo = unquote(storage_path.split("/")[-1])
                if not nome_arquivo:
                    nome_arquivo = f"fatura_{index}"

                archive.writestr(storage_path, documento_bytes)
    except HTTPException:
        raise
    except Exception:
        logger.exception("Erro ao gerar arquivo ZIP de faturas")
        raise HTTPException(
            status_code=500,
            detail="Erro interno ao gerar o ficheiro ZIP.",
        )

    zip_buffer.seek(0)
    filename = f"faturas_{data_inicio.strftime('%Y%m%d')}_{data_fim.strftime('%Y%m%d')}.zip"

    return StreamingResponse(
        content=zip_buffer,
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


@router.get("/zip")
async def gerar_zip(
    data_inicio: date = Query(..., description="Data de início (YYYY-MM-DD)"),
    data_fim: date = Query(..., description="Data de fim (YYYY-MM-DD)"),
):
    logger.info("Recebida requisição GET /zip para gerar pacote de faturas no período: %s a %s", data_inicio, data_fim)
    return _build_zip_response(data_inicio, data_fim)
